import os
import datetime
import dateutil.parser
import requests
import time
import argparse
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, fills
from openpyxl.utils import get_column_letter
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from datetime import datetime
from datetime import timedelta

parser = argparse.ArgumentParser()
parser.add_argument('apiKey')
parser.add_argument('testMode')
args = parser.parse_args()

apiKey = args.apiKey
apiURL = 'https://api.squarespace.com/1.0/commerce/orders'

productNameIgnoreList = ['Second Instalment', 'Spring 2019', '2019 - Beaumont Coding Classes', 'Stratford', 'Muslim National School', 'Custom payment amount']

def StringToBoolean(s):
    if s == 'True':
         return True
    elif s == 'False':
         return False

def GetDateTimeFromISO8601String(s):
    return dateutil.parser.parse(s)

def WriteLastGenerationDate(endDate):
    with open(os.path.dirname(os.path.abspath(__file__)) + '\LastClassListGenerationDate.txt', 'w') as file:
        file.write(endDate)

def ReadLastGenerationDate():
    with open(os.path.dirname(os.path.abspath(__file__)) + '\LastClassListGenerationDate.txt') as file:
        lastEndDate = GetDateTimeFromISO8601String(file.readlines()[0])
    startDate = lastEndDate + timedelta(microseconds=1)

    return startDate.strftime('%Y-%m-%dT%H:%M:%S.%fZ')

def ExportAllOrders():
    maxRetries = 10
    sleepTime = 1
    pageNumber = 1
    orderList = []
    responsePageList = []

    startDate = ReadLastGenerationDate()
    endDate = datetime.now().strftime('%Y-%m-%dT%H:%M:%S.%fZ')

    headers = {
        'Authorization': 'Bearer ' + apiKey,
    }

    params = [
        ('modifiedAfter', startDate),
        ('modifiedBefore', endDate)
    ]

    for i in range(maxRetries):
        response = requests.get(apiURL, headers=headers, params=params)

        if response.status_code == 200:
            responseJSON = response.json()
            break
        elif response.status_code == 429:
            print('Retrying...')
            time.sleep(sleepTime)
            sleepTime += 0.5
        else:
            print('Error getting request (Status Code: ' + str(response.status_code) + ')')
            print('Exiting program.')
            exit(-1)

    if response.status_code != 200:
        print('Error getting request (Status Code: ' + str(response.status_code) + ')')
        print('Exiting program.')
        exit(-1)

    for index in range(len(responseJSON['result'])):
        skipOrder = False
        for productName in productNameIgnoreList:
            if productName in responseJSON['result'][index]['lineItems'][0]['productName']:
                skipOrder = True
                break
        if not skipOrder:
            orderList.append(responseJSON['result'][index])

    print('Page ' + str(pageNumber) + ' Request Completed.')

    if orderList:
        responsePageList.append(orderList)

    while responseJSON['pagination']['hasNextPage'] == True:
        pageNumber += 1

        orderList = []

        params = [
            ('cursor', responseJSON['pagination']['nextPageCursor'])
        ]

        for i in range(maxRetries):
            response = requests.get(apiURL, headers=headers, params=params)

            if response.status_code == 200:
                responseJSON = response.json()
                break
            elif response.status_code == 429:
                print('Retrying...')
                time.sleep(sleepTime)
                sleepTime += 0.5
            else:
                print('Error getting request (Status Code: ' + str(response.status_code) + ')')
                print('Exiting program.')
                exit(-1)

        if response.status_code != 200:
            print('Error getting request (Status Code: ' + str(response.status_code) + ')')
            print('Exiting program.')
            exit(-1)

        for index in range(len(responseJSON['result'])):
            skipOrder = False
            for productName in productNameIgnoreList:
                if productName in responseJSON['result'][index]['lineItems'][0]['productName']:
                    skipOrder = True
                    break
            if not skipOrder:
                orderList.append(responseJSON['result'][index])

        print('Page ' + str(pageNumber) + ' Request Completed.')

        if orderList:
            responsePageList.append(orderList)

    return responsePageList, endDate

def ExportIndividualOrders(allOrdersList, testMode):
    maxRetries = 10
    sleepTime = 1
    orderCount = 1
    responseList = []
    classTypeList = []
    fullYearList = []

    headers = {
        'Authorization': 'Bearer ' + apiKey,
        'User-Agent' : 	'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36'
    }

    print('Exporting Individual Orders...')
    for orderList in allOrdersList:
        for order in orderList:
            orderID = order['id']

            for i in range(maxRetries):
                response = requests.get(apiURL + '/' + orderID, headers=headers)

                if response.status_code == 200:
                    responseJSON = response.json()
                    break
                elif response.status_code == 429:
                    print('Retrying...')
                    time.sleep(sleepTime)
                    sleepTime += 0.5
                else:
                    print('Error getting request (Status Code: ' + str(response.status_code) + ')')
                    print('Exiting program.')
                    exit(-1)

            if response.status_code != 200:
                print('Error getting request (Status Code: ' + str(response.status_code) + ')')
                print('Exiting program.')
                exit(-1)

            if 'Test' in responseJSON['lineItems'][0]['customizations'][0]['value']:
                if testMode:
                    if 'Tech Club' in responseJSON['lineItems'][0]['productName'].split('- ')[1]:
                        classType = (responseJSON['lineItems'][0]['productName'].split()[0],responseJSON['lineItems'][0]['productName'].split()[1], 'Tech Club')
                    else:
                        classType = (responseJSON['lineItems'][0]['productName'].split()[0],responseJSON['lineItems'][0]['productName'].split()[1])

                    for variantOption in responseJSON['lineItems'][0]['variantOptions']:
                        if 'Payment Plan' in variantOption['optionName']:
                            if '30 Weeks' in variantOption['value']:
                                classType = ('Spring', str(int(responseJSON['lineItems'][0]['productName'].split()[1]) + 1), 'Next Term')
                                fullYearList.append(responseJSON)
                                break

                    if classType not in classTypeList:
                        classTypeList.append(classType)

                    responseList.append(responseJSON)
                    print('Order ' + str(orderCount) + ' Completed.')
                    orderCount += 1
            else:
                if 'Tech Club' in responseJSON['lineItems'][0]['productName'].split('- ')[1]:
                    classType = (responseJSON['lineItems'][0]['productName'].split()[0],
                                 responseJSON['lineItems'][0]['productName'].split()[1], 'Tech Club')
                else:
                    classType = (responseJSON['lineItems'][0]['productName'].split()[0],
                                 responseJSON['lineItems'][0]['productName'].split()[1])

                for variantOption in responseJSON['lineItems'][0]['variantOptions']:
                    if 'Payment Plan' in variantOption['optionName']:
                        if '30 Weeks' in variantOption['value']:
                            classType = ('Spring', str(int(responseJSON['lineItems'][0]['productName'].split()[1]) + 1), 'Next Term')
                            fullYearList.append(responseJSON)
                            break

                if classType not in classTypeList:
                    classTypeList.append(classType)

                responseList.append(responseJSON)
                print('Order ' + str(orderCount) + ' Completed.')
                orderCount += 1

    return responseList, classTypeList, fullYearList

def GoogleDriveAccess():
    print('Getting Access to Google Drive.')

    gauth = GoogleAuth()
    gauth.LoadCredentialsFile("credentials.txt")

    if gauth.credentials is None:
        gauth.LocalWebserverAuth()
    elif gauth.access_token_expired:
        gauth.Refresh()
    else:
        gauth.Authorize()
    gauth.SaveCredentialsFile(os.path.dirname(os.path.abspath(__file__)) + "credentials.txt")

    return GoogleDrive(gauth)

def DownloadClassListsFromGoogleDrive(drive, classListName):
    downloadMimeType = None

    mimeTypes = {
        'application/vnd.google-apps.document': 'application/pdf',
        'application/vnd.google-apps.spreadsheet': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }

    if classListName + '.xlsx' in os.listdir("."):
        os.remove(classListName + '.xlsx')

    fileList = drive.ListFile({'q': "'root' in parents and trashed=false"}).GetList()

    for file in fileList:
        if 'Class lists' in file['title']:
            folderContentList = drive.ListFile({'q': "'{}' in parents and trashed=false".format(file['id'])}).GetList()

    for file in folderContentList:
        if classListName in file['title']:
            if file['mimeType'] in mimeTypes:
                print('Downloading Class Lists from Google Drive.')
                downloadMimeType = mimeTypes[file['mimeType']]
                file.GetContentFile(file['title'], mimetype=downloadMimeType)
            break

    if classListName not in os.listdir("."):
        wb = openpyxl.Workbook()
        wb.save(classListName + '.xlsx')

def AppendDfToExcel(fileName, df, sheetName, book, startRow=None, truncateSheet=False, **toExcelKwargs):
    if 'engine' in toExcelKwargs:
        toExcelKwargs.pop('engine')

    writer = pd.ExcelWriter(fileName, engine='openpyxl')

    try:
        writer.book = book

        if startRow is None and sheetName in writer.book.sheetnames:
            startRow = writer.book[sheetName].max_row

        if truncateSheet and sheetName in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheetName)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheetName, idx)

        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass

    if startRow is None:
        startRow = 0

    df.to_excel(writer, sheetName, header=False, index=False, startrow=startRow, **toExcelKwargs)

    writer.save()

def CreateAndAppendClassLists(orderList, classListName):
    print('Sorting Class Lists')

    for fileName in os.listdir(os.path.dirname(os.path.abspath(__file__))):
        if classListName in fileName:
            if '.xlsx' not in fileName:
                os.rename(fileName, fileName + '.xlsx')
                fileName = fileName + '.xlsx'

            seperatedNameOrderList = []

            for order in orderList:
                if len(order['lineItems'][0]['customizations'][0]['value'].split()) > 1 and len(
                        order['lineItems'][0]['customizations'][1]['value'].split()) > 1 and len(
                    order['lineItems'][0]['customizations'][2]['value'].split()) == 0:
                    seperatedNameOrderList.append(order)
                elif len(order['lineItems'][0]['customizations'][0]['value'].split()) > 1 and len(
                        order['lineItems'][0]['customizations'][1]['value'].split()) == 0 and len(
                    order['lineItems'][0]['customizations'][2]['value'].split()) > 1:
                    seperatedNameOrderList.append(order)
                elif len(order['lineItems'][0]['customizations'][0]['value'].split()) > 1 and len(
                        order['lineItems'][0]['customizations'][1]['value'].split()) > 1 and len(
                    order['lineItems'][0]['customizations'][2]['value'].split()) > 1:
                    seperatedNameOrderList.append(order)
                    seperatedNameOrderList.append(order)
                seperatedNameOrderList.append(order)

            prevOrderID = 0
            thirdOrder = False

            book = load_workbook(fileName)
            writer = pd.ExcelWriter(fileName, engine='openpyxl')
            writer.book = book

            if 'Summary' not in book.sheetnames:
                book.create_sheet("Summary")
                book.save(fileName)

            if 'Sheet' in book.sheetnames:
                sheet = book.get_sheet_by_name('Sheet')
                book.remove_sheet(sheet)

            classDates = []

            for order in seperatedNameOrderList:
                if order['orderNumber'] == prevOrderID:
                    if thirdOrder:
                        order['lineItems'][0]['customizations'][0]['value'] = order['lineItems'][0]['customizations'][2]['value']
                        order['lineItems'][0]['customizations'][3]['value'] = order['lineItems'][0]['customizations'][5]['value']
                        thirdOrder = False
                    else:
                        order['lineItems'][0]['customizations'][0]['value'] = order['lineItems'][0]['customizations'][1]['value']
                        order['lineItems'][0]['customizations'][3]['value'] = order['lineItems'][0]['customizations'][4]['value']
                        thirdOrder = True
                else:
                    thirdOrder = False

                prevOrderID = order['orderNumber']

                time = order['lineItems'][0]['variantOptions'][0]['value'].split(',')[1].replace(' ', '')[0:5].replace(':', '')

                if 'Tech Club' in fileName:
                    venue = order['lineItems'][0]['productName'].split('- ')[2].split(',')[0].split()[0]
                else:
                    venue = order['lineItems'][0]['productName'].split('- ')[1].split(',')[0].split()[0]

                if 'Summer' in fileName or 'Easter' in fileName:
                    day = order['lineItems'][0]['variantOptions'][0]['value'].split(',')[0].split()[1].split('-')[0]
                    month = order['lineItems'][0]['variantOptions'][0]['value'].split(',')[0].split()[0][0:3]
                    startDate = datetime.strptime(order['lineItems'][0]['variantOptions'][0]['value'].split(',')[0].split()[1].split('-')[0] + ' ' + order['lineItems'][0]['variantOptions'][0]['value'].split(',')[0].split()[0], '%d %B')
                    numClasses = 5
                    sheetName = venue.capitalize() + '_' + day + '_' + month + '_' + time
                elif 'Autumn' in fileName:
                    day = order['lineItems'][0]['variantOptions'][0]['value'][0:3]
                    startDate = ''
                    numClasses = 12
                    sheetName = day + '_' + venue.capitalize() + '_' + time
                elif 'Spring' in fileName:
                    day = order['lineItems'][0]['variantOptions'][0]['value'][0:3]
                    startDate = ''
                    numClasses = 18
                    sheetName = day + '_' + venue.capitalize() + '_' + time

                for i in range(0, numClasses):
                    classDates.append((startDate + timedelta(days=i)).strftime('%d %B'))

                if sheetName not in book.sheetnames:
                    writer = pd.ExcelWriter(fileName, engine='openpyxl')
                    writer.book = book

                    dataGapSize = 40

                    classListTemplateLines = [
                                              ['Attendence'],
                                              [''],
                                              ['Total No. Of Students'],
                                              [''],
                                              ['Date'],
                                              ['Student Name', 'Tutor'],
                                              ['Week ' + str(classNum + 1) for classNum in range(numClasses)],
                                              [' ' for gap in range(dataGapSize - 1)],
                                              ['Gender',
                                               'Order ID',
                                               'Email',
                                               'Billing Name',
                                               'Phone',
                                               'Student Name(s)',
                                               'Student Date(s) of Birth',
                                               'Student\'s School and Class',
                                               'Are you a returning Academy of Code student?',
                                               'If you are a returning student, when was your last term with AoC?',
                                               'Additional support for your child',
                                               'Photography Consent',
                                               'Other Details',
                                               'Other Teacher Notes']
                                              ]

                    classListTemplateLines[5] = classListTemplateLines[5] + classListTemplateLines[6] + classListTemplateLines[7] + classListTemplateLines[8]

                    for i in range(0,3):
                        del classListTemplateLines[6]

                    classListTemplateLines[4] = classListTemplateLines[4]+ [''] + classDates

                    templateDataFrame = pd.DataFrame(data=classListTemplateLines)
                    templateDataFrame.to_excel(writer, sheet_name=sheetName, header=False, index=False)

                    writer.save()

                if len(order['lineItems'][0]['customizations'][0]['value'].split()) == 1 and len(order['lineItems'][0]['customizations'][1]['value'].split()) == 1:
                    if len(order['lineItems'][0]['customizations'][2]['value'].split()) == 1:
                        d = [[order['lineItems'][0]['customizations'][0]['value'] + ' ' +
                              order['lineItems'][0]['customizations'][1]['value'] + ' ' +
                              order['lineItems'][0]['customizations'][2]['value'],
                              order['orderNumber'],
                              order['customerEmail'],
                              order['billingAddress']['firstName'] + ' ' +
                              order['billingAddress']['lastName'],
                              order['billingAddress']['phone'],
                              order['lineItems'][0]['customizations'][0]['value'] + ' ' +
                              order['lineItems'][0]['customizations'][1]['value'] + ' ' +
                              order['lineItems'][0]['customizations'][2]['value'],
                              order['lineItems'][0]['customizations'][3]['value'],
                              order['lineItems'][0]['customizations'][6]['value'],
                              order['lineItems'][0]['customizations'][7]['value'],
                              order['lineItems'][0]['customizations'][8]['value'],
                              order['lineItems'][0]['customizations'][10]['value'],
                              order['lineItems'][0]['customizations'][11]['value'],
                              '']
                             ]
                    else:
                        d = [[order['lineItems'][0]['customizations'][0]['value'] + ' ' +
                              order['lineItems'][0]['customizations'][1]['value'],
                              order['orderNumber'],
                              order['customerEmail'],
                              order['billingAddress']['firstName'] + ' ' +
                              order['billingAddress']['lastName'],
                              order['billingAddress']['phone'],
                              order['lineItems'][0]['customizations'][0]['value'] + ' ' +
                              order['lineItems'][0]['customizations'][1]['value'],
                              order['lineItems'][0]['customizations'][3]['value'],
                              order['lineItems'][0]['customizations'][6]['value'],
                              order['lineItems'][0]['customizations'][7]['value'],
                              order['lineItems'][0]['customizations'][8]['value'],
                              order['lineItems'][0]['customizations'][10]['value'],
                              order['lineItems'][0]['customizations'][11]['value'],
                              '']
                            ]
                elif len(order['lineItems'][0]['customizations'][0]['value'].split()) == 1 and len(order['lineItems'][0]['customizations'][2]['value'].split()) == 1:
                    if len(order['lineItems'][0]['customizations'][1]['value'].split()) == 1:
                        d = [[order['lineItems'][0]['customizations'][0]['value'] + ' ' +
                              order['lineItems'][0]['customizations'][1]['value'] + ' ' +
                              order['lineItems'][0]['customizations'][2]['value'],
                              order['orderNumber'],
                              order['customerEmail'],
                              order['billingAddress']['firstName'] + ' ' +
                              order['billingAddress']['lastName'],
                              order['billingAddress']['phone'],
                              order['lineItems'][0]['customizations'][0]['value'] + ' ' +
                              order['lineItems'][0]['customizations'][1]['value'] + ' ' +
                              order['lineItems'][0]['customizations'][2]['value'],
                              order['lineItems'][0]['customizations'][3]['value'],
                              order['lineItems'][0]['customizations'][6]['value'],
                              order['lineItems'][0]['customizations'][7]['value'],
                              order['lineItems'][0]['customizations'][8]['value'],
                              order['lineItems'][0]['customizations'][10]['value'],
                              order['lineItems'][0]['customizations'][11]['value'],
                              '']
                             ]
                    else:
                        d = [[order['lineItems'][0]['customizations'][0]['value'] + ' ' +
                              order['lineItems'][0]['customizations'][2]['value'],
                              order['orderNumber'],
                              order['customerEmail'],
                              order['billingAddress']['firstName'] + ' ' +
                              order['billingAddress']['lastName'],
                              order['billingAddress']['phone'],
                              order['lineItems'][0]['customizations'][0]['value'] + ' ' +
                              order['lineItems'][0]['customizations'][2]['value'],
                              order['lineItems'][0]['customizations'][3]['value'],
                              order['lineItems'][0]['customizations'][6]['value'],
                              order['lineItems'][0]['customizations'][7]['value'],
                              order['lineItems'][0]['customizations'][8]['value'],
                              order['lineItems'][0]['customizations'][10]['value'],
                              order['lineItems'][0]['customizations'][11]['value'],
                              '']
                             ]
                else:
                    d = [[order['lineItems'][0]['customizations'][0]['value'],
                          order['orderNumber'],
                          order['customerEmail'],
                          order['billingAddress']['firstName'] + ' ' +
                          order['billingAddress']['lastName'],
                          order['billingAddress']['phone'],
                          order['lineItems'][0]['customizations'][0]['value'],
                          order['lineItems'][0]['customizations'][3]['value'],
                          order['lineItems'][0]['customizations'][6]['value'],
                          order['lineItems'][0]['customizations'][7]['value'],
                          order['lineItems'][0]['customizations'][8]['value'],
                          order['lineItems'][0]['customizations'][10]['value'],
                          order['lineItems'][0]['customizations'][11]['value'],
                          '']
                         ]

                book.save(fileName)

                book = load_workbook(fileName)

                maxCol = book.get_sheet_by_name(sheetName).max_column

                for k in range(maxCol-len(d[0])-1):
                    d[0].insert(1, '')

                orderDataFrame = pd.DataFrame(data=d)

                AppendDfToExcel(fileName, orderDataFrame, sheetName, book)

                totalStudentsCell = book[sheetName].cell(row=3, column=2)
                totalStudentsCell.value = book[sheetName].max_row - 6

                for i in range(1, len(classDates) + 2):
                    attendanceCell = book[sheetName].cell(row=1, column=i)
                    if i > 1:
                        attendanceCell.value = '=COUNTIF(' + get_column_letter(i) + '7:' + get_column_letter(i) + str(book[sheetName].max_row) + ',"Y")'
                    attendanceCell.fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb='00FFF2CC'))

                for row in range(1, 7):
                    for col in range(1, book[sheetName].max_column + 1):
                        colLetter = get_column_letter(col)
                        cell = book[sheetName][colLetter + str(row)]
                        cell.font = Font(bold=True)

                classDates = []

                writer.save()
                book.save(fileName)

            SortWorkSheets(fileName)

def SortWorkSheets(fileName):
    dayList = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun']
    seperateDayList = []
    newOrder = []

    book = load_workbook(fileName)

    if 'Summer' not in fileName and 'Easter' not in fileName:
        for day in dayList:
            sheetList = []
            for sheet in book.sheetnames:
                if day == sheet[0:3]:
                    sheetList.append(sheet)
            if sheetList:
                seperateDayList.append(sheetList)

        for day in seperateDayList:
            day.sort()
        newClassOrderList = [j for i in seperateDayList for j in i]

        for sheetName in newClassOrderList:
            newOrder.append(book.worksheets.index(book.get_sheet_by_name(sheetName)))

        summarySheetIndex = book.worksheets.index(book.get_sheet_by_name('Summary'))
    else:
        book._sheets.sort(key=lambda ws: ws.title)
        summarySheetIndex = book.worksheets.index(book.get_sheet_by_name('Summary'))
        newOrder = [i for i in range(len(book.worksheets))]
        del newOrder[summarySheetIndex]

    newOrder.insert(0, summarySheetIndex)
    book._sheets = [book._sheets[i] for i in newOrder]
    book.save(fileName)

def DeleteOldFileFromGoogleDrive(drive, classListName):
    fileList = drive.ListFile({'q': "'root' in parents and trashed=false"}).GetList()

    for file in fileList:
        if 'Class lists' in file['title']:
            folderContentList = drive.ListFile({'q': "'{}' in parents and trashed=false".format(file['id'])}).GetList()

    for file in folderContentList:
        if classListName in file['title']:
            file.Delete()
            print('Deleting ' + file['title'] + ' from Google Drive.')
            break

def UploadToGoogleDrive(drive, classListName):
    fileList = drive.ListFile({'q': "'root' in parents and trashed=false"}).GetList()

    for file in fileList:
        if 'Class lists' in file['title']:
            folderId = file['id']
            fileMetaData = {'title': classListName + '.xlsx', "parents": [{"id": folderId, "kind": "drive#childList"}]}
            folder = drive.CreateFile(fileMetaData)
            folder.SetContentFile(classListName + '.xlsx')
            print('Uploading ' + file['title'] + ' from Google Drive.')
            folder.Upload({'convert': True})
            break

def main():
    start = time.time()

    testMode = StringToBoolean(args.testMode)

    allOrdersList, endDate = ExportAllOrders()
    individualOrdersList, classTypeList, fullYearList = ExportIndividualOrders(allOrdersList, testMode)

    if len(individualOrdersList) > 0:
        drive = GoogleDriveAccess()
        for classType in classTypeList:
            if classType[0] == 'Summer':
                classListName = 'Summer Camps ' + classType[1]
            elif classType[0] == 'Easter':
                classListName = 'Easter Camps ' + classType[1]
            else:
                if len(classType) > 2 and classType[2] == 'Tech Club':
                    classListName = classType[2] + ' ' + classType[0] + ' ' + classType[1]
                else:
                    classListName = 'Evening&Weekends ' + classType[0] + ' ' + classType[1]

            if testMode:
                classListName = 'Test ' + classListName

            DownloadClassListsFromGoogleDrive(drive, classListName)

            if len(classType) > 2 and classType[2] == 'Next Term':
                CreateAndAppendClassLists(fullYearList, classListName)
            else:
                CreateAndAppendClassLists(individualOrdersList, classListName)
            DeleteOldFileFromGoogleDrive(drive, classListName)
            UploadToGoogleDrive(drive, classListName)

    WriteLastGenerationDate(endDate)

    end = time.time()

    print(str(round(end - start, 2)) + ' secs')

if __name__ == '__main__':
    main()